import Contract
import openpyxl
import os
from datetime import datetime
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment


thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

dateToday = datetime.today().strftime('%Y-%m-%d')
year = dateToday[0:4]
month = dateToday[5:7]
#get date for today for later use


def checkNewSheets(path):
    path = path + '\\' + '销售统计'
    os.chdir(path)
    entries = os.listdir()
    setOfEntries = set(entries)
    #change directory to the sale record folder

    fileName = year + '销售统计' + '.xlsx'
    #make sure there all sheets are already created, if not create new ones
    if fileName in setOfEntries:


        wb = openpyxl.load_workbook(fileName)
        sheets = wb.get_sheet_names()
        if (month + '月') not in sheets:
            wb.create_sheet(title=month + '月')
            pageSetup(wb, (month + '月'))
            wb.save(year + '销售统计' + '.xlsx')
    else:
        wb = openpyxl.Workbook()
        sheet  = wb.active
        sheet.title = month + '月'
        pageSetup(wb, (month + '月'))
        wb.save(year + '销售统计' + '.xlsx')



def enterData(path, listContract):
    path = path + '\\' + '销售统计'
    os.chdir(path)
    wb = openpyxl.load_workbook(year + '销售统计' + '.xlsx')
    sheet = wb.get_sheet_by_name(month + '月')

    #the following for loop finds the closest empty row in the sheet
    for i in sheet['A']:
        if i.value == None:
            max = int(i.row)
            break
        else:
            max = sheet.max_row + 1
    # keep count of each individual product within each entry
    DCA  = 0
    YC = 0
    RV = 0
    YCX = 0
    YCM = 0
    YCS = 0
    DD = 0

    for contract in listContract:
        for i in range(len(contract.price)):
            sheet.cell(row=max, column=1).value = dateToday
            sheet.cell(row=max, column=1).border = thin_border
            sheet.cell(row=max, column=2).value = contract.contractNum
            sheet.cell(row=max, column=2).border = thin_border
            sheet.cell(row=max, column=3).value = contract.companyName
            sheet.cell(row=max, column=3).border = thin_border
            sheet.cell(row=max, column=4).value = contract.modelNumber[i]
            sheet.cell(row=max, column=4).border = thin_border
            sheet.cell(row=max, column=4).alignment = Alignment(horizontal='left')
            sheet.cell(row=max, column=5).value = contract.modelCount[i]
            sheet.cell(row=max, column=5).border = thin_border
            sheet.cell(row=max, column=5).alignment = Alignment(horizontal='center')
            sheet.cell(row=max, column=6).value = contract.getFormattedPrice()[i]
            sheet.cell(row=max, column=6).border = thin_border
            sheet.cell(row=1, column=6).value += contract.price[i] * contract.modelCount[i]
            if 'DCA' in contract.modelNumber[i]:
                DCA += contract.modelCount[i]
            elif 'YC' in contract.modelNumber[i] and 'YCX' not in contract.modelNumber[i] and 'YCX' not in contract.modelNumber[i] and 'YCS' not in contract.modelNumber[i] :
                YC += contract.modelCount[i]
            elif 'RV' in contract.modelNumber[i]:
                RV += contract.modelCount[i]
            elif 'YCX' in contract.modelNumber[i] and 'YCM' not in contract.modelNumber[i] and 'YCS' not in contract.modelNumber[i]:
                YCX += contract.modelCount[i]
            elif 'YCM' in contract.modelNumber[i] and 'YCS' not in contract.modelNumber[i]:
                YCM += contract.modelCount[i]
            elif 'YCS' in contract.modelNumber[i]:
                YCS += contract.modelCount[i]
            elif 'DD' in contract.modelNumber[i]:
                DD += contract.modelCount[i]
            max += 1
    sheet['H3'] = sheet['H3'].value + DCA
    sheet['H4'] = sheet['H4'].value + YC
    sheet['H5'] = sheet['H5'].value + YCX
    sheet['H6'] = sheet['H6'].value + YCM
    sheet['H7'] = sheet['H7'].value + YCS
    sheet['H8'] = sheet['H8'].value + DD
    sheet['H9'] = sheet['H9'].value + RV
    sheet['H2'] = sheet['H3'].value + sheet['H4'].value + sheet['H5'].value + sheet['H6'].value + sheet['H7'].value + sheet['H8'].value + sheet['H9'].value

    wb.save(year + '销售统计' + '.xlsx')

def pageSetup(wb, sheetName):
    #standard setup for a new sheet
    sheet = wb.get_sheet_by_name(sheetName)
    sheet.freeze_panes = 'A3'
    sheet['A1'].font = Font(name='Times New Roman', bold=True)
    sheet['A2'].font = Font(name='Times New Roman', bold=True)
    sheet['B2'].font = Font(name='Times New Roman', bold=True)
    sheet['C2'].font = Font(name='Times New Roman', bold=True)
    sheet['D2'].font = Font(name='Times New Roman', bold=True)
    sheet['E2'].font = Font(name='Times New Roman', bold=True)
    sheet['F2'].font = Font(name='Times New Roman', bold=True)
    sheet['F1'].number_format = '#,##0 ¥'
    sheet['A1'] = year + '年' + month + '月销售合同清单'
    sheet['A2'] = '日期'
    sheet['B2'] = '合同号'
    sheet['C2'] = '订货单位信息'
    sheet['D2'] = '型号'
    sheet['E1'] = '总成交额：'
    sheet['F1'] = 0
    sheet['E2'] = '台数'
    sheet['F2'] = '价格'
    default_cells = ['A1', 'A2', 'B2', 'C2', 'D2', 'E2', 'F2', 'G1', 'G2', 'G3', 'G4', 'G5', 'G6', 'G7', 'G8', 'G9',
                     'H2', 'H3', 'H4', 'H5', 'H6', 'H7', 'H8', 'H9']
    sheet['G2'] = '总销量: '
    sheet['G3'] = 'DCA 销量：'
    sheet['G4'] = 'YC 销量：'
    sheet['G5'] = 'YCX 销量：'
    sheet['G6'] = 'YCM 销量：'
    sheet['G7'] = 'YCS 销量：'
    sheet['G8'] = 'DD 销量：'
    sheet['G9'] = 'RV 销量：'
    sheet['H3'] = 0
    sheet['H4'] = 0
    sheet['H5'] = 0
    sheet['H6'] = 0
    sheet['H7'] = 0
    sheet['H8'] = 0
    sheet['H9'] = 0
    for c in default_cells:
        sheet[c].border = thin_border
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 30
    sheet.column_dimensions['F'].width = 20
    sheet.column_dimensions['E'].width = 10
    sheet.column_dimensions['A'].width = 15

def makeContractList(path):
    path = path + '\\' + '前台拷贝'
    os.chdir(path)
    listOfFile = os.listdir()
    dict = {}
    for i in range(len(listOfFile)):
        dict["contract{0}".format(i)] = Contract.Contract(path + '\\' + listOfFile[i])

    x = dict.values()
    listOfContracts = []
    for i in x:
        listOfContracts.append(i)
    return listOfContracts



