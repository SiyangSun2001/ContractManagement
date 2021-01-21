import openpyxl
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
import os

#formate for thin border in excel
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
#format date and time for time stamp in excel
from datetime import datetime
dateToday = datetime.today().strftime('%Y-%m-%d')

#creates a new work book and enters the data from contract obj
def createNewWBandEnter(path, Contract):
    os.chdir(path + '\\' + '客户信息')
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = Contract.companyName + '信息'
    sheet.freeze_panes = 'A2'
    sheet['A1'].font = Font(name='Times New Roman', bold=True)
    sheet['A1'] = '已 订 货 用 户 情 况 记 录 单'
    sheet['A2'] = '单位名称：'
    sheet['A3'] = '地址：'
    sheet['A4'] = '电话：'
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 30
    sheet.column_dimensions['F'].width = 30
    sheet['A2'] = sheet['A2'].value + Contract.companyFullName
    sheet['A4'] = sheet['A4'].value + Contract.phone
    sheet['A3'] = sheet['A3'].value + Contract.address
    sheet['A5'] = '次数'
    sheet['B5'] = '日期'
    sheet['C5'] = '合同编号'
    sheet['D5'] = '型号'
    sheet['E5'] = '台数'
    sheet['F5'] = '价格'
    default_Cell = ['A5', 'B5', 'C5', 'D5', 'E5', 'F5']
    for c in default_Cell:
        sheet[c].border = thin_border
    max = sheet.max_row
    for i in range(len(Contract.price)):
        sheet.cell(row=max + 1, column=1).value = max - 4
        sheet.cell(row=max + 1, column=2).value = dateToday
        sheet.cell(row=max + 1, column=3).value = Contract.contractNum
        sheet.cell(row=max + 1, column=1).alignment = Alignment(horizontal='center')
        sheet.cell(row=max + 1, column=2).alignment = Alignment(vertical='top')
        sheet.cell(row=max + 1, column=3).alignment = Alignment(horizontal='center')
        sheet.cell(row=max + 1, column=1).border = thin_border
        sheet.cell(row=max + 1, column=2).border = thin_border
        sheet.cell(row=max + 1, column=3).border = thin_border
        sheet.cell(row=max + 1, column=4).value = Contract.modelNumber[i]
        sheet.cell(row=max + 1, column=4).alignment = Alignment(vertical='top')
        sheet.cell(row=max + 1, column=4).border = thin_border

        sheet.cell(row=max + 1, column=5).value = Contract.modelCount[i]
        sheet.cell(row=max + 1, column=5).alignment = Alignment(horizontal='center')
        sheet.cell(row=max + 1, column=5).border = thin_border

        sheet.cell(row=max + 1, column=6).value = Contract.getFormattedPrice()[i]
        sheet.cell(row=max + 1, column=6).alignment = Alignment(vertical='top')
        sheet.cell(row=max + 1, column=6).border = thin_border
    wb.save('已订货客户' + ' ' + Contract.companyName + '.xlsx')

#enter data into a workbook from a contract obj
def enterCustomerInfo(path, Contract):
    path = path + '\\' + '客户信息'
    os.chdir(path)
    fileName = '已订货客户' + ' ' + Contract.companyName + '.xlsx'
    wb = openpyxl.load_workbook(fileName)
    sheet = wb.active
    max = sheet.max_row
    for i in range(len(Contract.price)):
        sheet.cell(row=max + 1, column=1).value = max - 4
        sheet.cell(row=max + 1, column=2).value = dateToday
        sheet.cell(row=max + 1, column=3).value = Contract.contractNum
        sheet.cell(row=max + 1, column=1).alignment = Alignment(horizontal='center')
        sheet.cell(row=max + 1, column=2).alignment = Alignment(vertical='top')
        sheet.cell(row=max + 1, column=3).alignment = Alignment(horizontal ='center')
        sheet.cell(row=max + 1, column=1).border = thin_border
        sheet.cell(row=max + 1, column=2).border = thin_border
        sheet.cell(row=max + 1, column=3).border = thin_border
        sheet.cell(row=max + 1, column=4).value = Contract.modelNumber[i]
        sheet.cell(row=max + 1, column=4).alignment = Alignment(vertical='top')
        sheet.cell(row=max + 1, column=4).border = thin_border

        sheet.cell(row=max + 1, column=5).value = Contract.modelCount[i]
        sheet.cell(row=max + 1, column=5).alignment = Alignment(horizontal ='center')
        sheet.cell(row=max + 1, column=5).border = thin_border

        sheet.cell(row=max + 1, column=6).value = Contract.getFormattedPrice()[i]
        sheet.cell(row=max + 1, column=6).alignment = Alignment(vertical='top')
        sheet.cell(row=max + 1, column=6).border = thin_border
    wb.save('已订货客户' + ' ' + Contract.companyName + '.xlsx')

